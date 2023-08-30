import time
import subprocess
import numpy
import numpy as np
import pandas as pd
import self as self
from openpyxl import load_workbook
import psutil
import signal
import tkinter as tk
import os
from pprint import pprint
from openpyxl import load_workbook
import math as math

# forced shutdown excel to complete saving
os.system("taskkill /f /im  EXCEL.exe")

# path to Analysis_Plan_Template.xlsx file
test_file = os.path.realpath('Iveco_Cluster_Ticket_List_xlsx.xlsx')
file_path_2 = os.path.realpath('Analysis_Plan_Iveco_Cluster.xlsx')
file_path = os.path.realpath('testowy_plik.xlsx')  ## plik testowy excela do uzupelnienia

# loading excel file to script and reading properly sheet
issues = pd.read_excel(test_file,sheet_name='Issues')
setup = pd.read_excel(file_path,sheet_name='Setup')
block_interface = pd.read_excel(file_path,sheet_name='Block | Interface')
block_type = pd.read_excel(file_path_2,sheet_name='Blocks')
program_dashboard = pd.read_excel(file_path_2,sheet_name='Program dashboard')
program_timing = pd.read_excel(file_path_2,sheet_name='Program timing')
other_info = pd.read_excel(file_path_2,sheet_name='Other info')
# print(program_dashboard.iat[3,2])
# print(other_info.iat[0,0])
# print(other_info)
# print(program_dashboard.iat[1,2] + " " + program_dashboard.iat[0,2])
# const names
contact_ecae = program_dashboard.iat[7, 2]
assigne = program_dashboard.iat[5, 2]
# start and end date
start = program_timing.iat[0,1]
end = program_timing.iat[0,2]
#
name_project = program_dashboard.iat[1,2] + " " + program_dashboard.iat[0,2]
story = "Story"
analysis_story = "Analysis/Simulation"
rev_story = "Analysis Review"
hasztag = other_info.iat[0,0]
prod_desc= program_dashboard.iat[1,2]
roster_num= program_dashboard.iat[1,2]
design_center = program_dashboard.iat[3,2]
# iat[kolumny od 0,wiersze od 0]
prog_phase= other_info.iat[0,1]
# print(prog_phase)
prog_compl = program_dashboard.iat[4,2]
samp_var = program_timing.iat[0,0]
prod_line = program_dashboard.iat[2,2]

# time to review and analysis
rev_time = 1
pi_sim_time= other_info.iat[2,6]
wcca_time = 16
si_time = 24
ac_time = 24
spar_time = 8
layout_time=1
tran_ce_time=1
tran_esd_time=1
tran_pul_time=1
spar_tdr_time=1
re_time=1
wcca_time=1

# Opening excel file , path to  executable file excel and xlsx file
prog = r"c:\Program Files\Microsoft Office\root\Office16\EXCEL.EXE" # path to exe file of excel , flexible path
# OpenIt = subprocess.Popen([prog, file])


# list_2 ma miec wielkosc ilosci analiz w excelu analysis plan template czyli ilosci wierszy
list_2 = np.empty([2,block_type.shape[0]],dtype=object)
optional = np.empty([2,block_type.shape[0]],dtype=object)
priority = np.empty([2,block_type.shape[0]],dtype=object)
block = np.empty([1,block_type.shape[0]],dtype=object)
comment = np.empty([1,block_type.shape[0]],dtype=object)
# wb = load_workbook(file_path,data_only=True)
# sheet_setup = wb['Setup']
wb_issues = load_workbook(test_file)
sheet_issues = wb_issues['Issues']


for l in range(0, block_type.shape[0]):
    list_2[0][l] = str(block_type[l:l + 1]["Block name"].values)  ## list of block type
    # print(list_2[0][l].replace('[', '').replace(']', '+'))
    for s in range(1,2):
        list_2[s][l] = block_type[l:l + 1]["Default analysis list"].str.split(', ', expand=True).values.tolist()
        optional[s][l] = block_type[l:l + 1]["Optional analysis list"].str.split(', ', expand=True).values.tolist()
        priority[s][l] = block_type[l:l + 1]["Priority"].str.split(', ', expand=True).values.tolist()
        block[0][l] = block_type[l:l + 1]["Block type"].str.split(', ', expand=True).values.tolist()
        comment[0][l] = block_type[l:l + 1]["Comments"].str.split(', ', expand=True).values

# print("komentarze -_>",comment)
priority[1] = [','.join(item[0]) for item in priority[1]]
# print("TUTAJ jestem->",list_2)
# print("projrytet->",priority[1])
# print("tutaj - > 2",list_2[0])
# print("tutaj",block[0])
# konwersja formatu
block[0] = [','.join(item[0]) for item in block[0]]
# comment[0] = [','.join(item[0]) for item in comment[0]]
# print(comment[0])
# print(np.shape(block[0]))

for kl in range(1, 101):
    # numeration of first column and all row
    sheet_issues.cell(row=kl + 1, column=1).value = kl
    # Status
    sheet_issues.cell(row=kl + 1, column=5).value = "New"
    # contact ECAE --> Author
    sheet_issues.cell(row=kl + 1, column=9).value = contact_ecae
    sheet_issues.cell(row=kl + 1, column=19).value = prog_phase
    sheet_issues.cell(row=kl + 1, column=20).value = prog_compl
    # start date and end date
    sheet_issues.cell(row=kl + 1, column=11).value = start
    sheet_issues.cell(row=kl + 1, column=12).value = end
    sheet_issues.cell(row=kl + 1, column=17).value = prod_line
    sheet_issues.cell(row=kl + 1, column=14).value = "0"

wb_issues.save(test_file)
print(comment)

row_index = 2
for i in range(len(list_2[0])):
    domain = list_2[0][i].strip("[]'")
    # print(list_2[0][i])
    analyses = list_2[1][i]
    sheet_issues.cell(row=row_index, column=2).value = name_project
    sheet_issues.cell(row=row_index, column=6).value = priority[1][i]
    sheet_issues.cell(row=row_index, column=7).value = "[" + samp_var + "]" + " " + domain
    if str(comment[0][i].item()) == 'nan':
        sheet_issues.cell(row=row_index, column=8).value = "To be filled by analysis team."
    else:
        sheet_issues.cell(row=row_index, column=8).value = str(comment[0][i].item())
    sheet_issues.cell(row=row_index, column=30).value= domain
    sheet_issues.cell(row=row_index, column=31).value = assigne
    sheet_issues.cell(row=row_index, column=35).value = samp_var
    sheet_issues.cell(row=row_index, column=38).value = block[0][i]
    sheet_issues.cell(row=row_index, column=3).value = story
    sheet_issues.cell(row=row_index, column=15).value = program_dashboard.iat[0,2]
    sheet_issues.cell(row=row_index, column=16).value = design_center
    sheet_issues.cell(row=row_index, column=18).value = roster_num
    sheet_issues.cell(row=row_index, column=29).value = roster_num
    sheet_issues.cell(row=row_index, column=10).value = assigne
    p= row_index -1
    sheet_issues.cell(row=row_index, column=4).value = hasztag
    row_index += 1
    for analysis in analyses:
        for element in analysis:
            sheet_issues.cell(row=row_index, column=2).value = name_project
            if element == 'PI':
                sheet_issues.cell(row=row_index, column=2).value = name_project
                sheet_issues.cell(row=row_index, column=4).value = p
                sheet_issues.cell(row=row_index, column=7).value = "[" + samp_var + "]" + " " + element + " AC " + list_2[0][i].strip("[]'")
                sheet_issues.cell(row=row_index, column=6).value = priority[1][i]
                sheet_issues.cell(row=row_index, column=26).value = "TBD"
                sheet_issues.cell(row=row_index, column=39).value = "NO"
                sheet_issues.cell(row=row_index, column=3).value = analysis_story
                sheet_issues.cell(row=row_index, column=13).value = pi_sim_time
                sheet_issues.cell(row=row_index, column=21).value = "HL_PI"
                row_index += 1
                sheet_issues.cell(row=row_index, column=2).value = name_project
                sheet_issues.cell(row=row_index, column=7).value = "[" + samp_var + "]" + " " + element + " AC " + list_2[0][i].strip("[]'") + " REVIEW"
                sheet_issues.cell(row=row_index, column=6).value = priority[1][i]
                sheet_issues.cell(row=row_index, column=34).value= "Teams / Skype"
                sheet_issues.cell(row=row_index, column=26).value = "TBD"
                sheet_issues.cell(row=row_index, column=4).value = (sheet_issues.cell(row=row_index,column=1).value) - 1
                sheet_issues.cell(row=row_index, column=3).value = rev_story
                sheet_issues.cell(row=row_index, column=13).value = rev_time
                sheet_issues.cell(row=row_index, column=21).value = "HL_PI"
                row_index += 1
                sheet_issues.cell(row=row_index, column=2).value = name_project
                sheet_issues.cell(row=row_index, column=4).value = p
                sheet_issues.cell(row=row_index, column=7).value = "[" + samp_var + "]" + " " + element + " DC " + list_2[0][i].strip("[]'")
                sheet_issues.cell(row=row_index, column=26).value = "TBD"
                sheet_issues.cell(row=row_index, column=6).value = priority[1][i]
                sheet_issues.cell(row=row_index, column=39).value = "NO"
                sheet_issues.cell(row=row_index, column=3).value = analysis_story
                sheet_issues.cell(row=row_index, column=13).value = pi_sim_time
                sheet_issues.cell(row=row_index, column=21).value = "HL_PI"
                row_index += 1
                sheet_issues.cell(row=row_index, column=2).value = name_project
                sheet_issues.cell(row=row_index, column=7).value = "[" + samp_var + "]" + " " + element + " DC " + list_2[0][i].strip("[]'") + " REVIEW"
                sheet_issues.cell(row=row_index, column=6).value = priority[1][i]
                sheet_issues.cell(row=row_index, column=34).value = "Teams / Skype"
                sheet_issues.cell(row=row_index, column=26).value = "TBD"
                sheet_issues.cell(row=row_index, column=4).value = (sheet_issues.cell(row=row_index, column=1).value) - 1
                sheet_issues.cell(row=row_index, column=3).value = rev_story
                sheet_issues.cell(row=row_index, column=13).value = rev_time
                sheet_issues.cell(row=row_index, column=21).value = "HL_PI"
                row_index += 1
            else:
                sheet_issues.cell(row=row_index, column=2).value = name_project
                sheet_issues.cell(row=row_index, column=4).value = p
                sheet_issues.cell(row=row_index, column=7).value = "[" + samp_var + "]" + " " + element + " " + list_2[0][i].strip("[]'")
                sheet_issues.cell(row=row_index, column=6).value = priority[1][i]
                sheet_issues.cell(row=row_index, column=39).value = "NO"
                sheet_issues.cell(row=row_index, column=26).value = "TBD"
                sheet_issues.cell(row=row_index, column=3).value = analysis_story
                # wykrycie rodzaju analizy i wpis czasu do excela
                if element == 'SI':
                    sheet_issues.cell(row=row_index, column=13).value = si_time
                    sheet_issues.cell(row=row_index, column=21).value = "tuul"
                    sheet_issues.cell(row=row_index+1, column=21).value = "tuul2"
                    (sheet_issues.cell(row=row_index, column=24).value,sheet_issues.cell(row=row_index + 1, column=24).value) = ("tajpAnalisys", "tajp2")
                elif element == 'AC/Stability':
                    sheet_issues.cell(row=row_index, column=13).value = ac_time
                    sheet_issues.cell(row=row_index, column=21).value = "tuul"
                    sheet_issues.cell(row=row_index+1, column=21).value = "tuul2"
                    (sheet_issues.cell(row=row_index, column=24).value,sheet_issues.cell(row=row_index + 1, column=24).value) = ("tajpAnalisys", "tajp2")
                elif element == 'Layout review':
                    sheet_issues.cell(row=row_index, column=13).value = layout_time
                    sheet_issues.cell(row=row_index, column=21).value = "tuul"
                    sheet_issues.cell(row=row_index+1, column=21).value = "tuul2"
                    (sheet_issues.cell(row=row_index, column=24).value,sheet_issues.cell(row=row_index + 1, column=24).value) = ("tajpAnalisys", "tajp2")
                elif element == 'Transient CE':
                    sheet_issues.cell(row=row_index, column=13).value = tran_ce_time
                    sheet_issues.cell(row=row_index, column=21).value = "tuul"
                    sheet_issues.cell(row=row_index+1, column=21).value = "tuul2"
                    (sheet_issues.cell(row=row_index, column=24).value,sheet_issues.cell(row=row_index + 1, column=24).value) = ("tajpAnalisys", "tajp2")
                elif element == 'Transient ESD':
                    sheet_issues.cell(row=row_index, column=13).value = tran_esd_time
                    sheet_issues.cell(row=row_index, column=21).value = "tuul"
                    sheet_issues.cell(row=row_index+1, column=21).value = "tuul2"
                    (sheet_issues.cell(row=row_index, column=24).value,sheet_issues.cell(row=row_index + 1, column=24).value) = ("tajpAnalisys", "tajp2")
                elif element == 'Transient pulses':
                    sheet_issues.cell(row=row_index, column=13).value = tran_pul_time
                    sheet_issues.cell(row=row_index, column=21).value = "tuul"
                    sheet_issues.cell(row=row_index+1, column=21).value = "tuul2"
                    (sheet_issues.cell(row=row_index, column=24).value,sheet_issues.cell(row=row_index + 1, column=24).value) = ("tajpAnalisys", "tajp2")
                elif element == 'S-Par/TDR':
                    sheet_issues.cell(row=row_index, column=13).value = spar_tdr_time
                    sheet_issues.cell(row=row_index, column=21).value = "tuul"
                    sheet_issues.cell(row=row_index+1, column=21).value = "tuul2"
                    (sheet_issues.cell(row=row_index, column=24).value,sheet_issues.cell(row=row_index + 1, column=24).value) = ("tajpAnalisys", "tajp2")
                elif element == 'RE':
                    sheet_issues.cell(row=row_index, column=13).value = re_time
                    sheet_issues.cell(row=row_index, column=21).value = "tuul"
                    sheet_issues.cell(row=row_index+1, column=21).value = "tuul2"
                    (sheet_issues.cell(row=row_index, column=24).value,sheet_issues.cell(row=row_index + 1, column=24).value) = ("tajpAnalisys", "tajp2")
                else:
                    sheet_issues.cell(row=row_index, column=13).value = wcca_time
                    sheet_issues.cell(row=row_index, column=21).value = "tuul"
                    sheet_issues.cell(row=row_index+1, column=21).value = "tuul2"
                    sheet_issues.cell(row=row_index, column=39).value = "NO"
                    (sheet_issues.cell(row=row_index, column=24).value,sheet_issues.cell(row=row_index + 1, column=24).value) = ("tajpAnalisys", "tajp2")
                row_index += 1
                sheet_issues.cell(row=row_index, column=2).value = name_project
                sheet_issues.cell(row=row_index, column=7).value = "[" + samp_var + "]" + " " + element + " " + list_2[0][i].strip("[]'") + " REVIEW"
                sheet_issues.cell(row=row_index, column=6).value = priority[1][i]
                sheet_issues.cell(row=row_index, column=34).value = "Teams / Skype"
                sheet_issues.cell(row=row_index, column=26).value = "TBD"
                sheet_issues.cell(row=row_index, column=4).value = (sheet_issues.cell(row=row_index,column=1).value) - 1
                sheet_issues.cell(row=row_index, column=3).value = rev_story
                sheet_issues.cell(row=row_index, column=13).value = rev_time
                row_index += 1



wb_issues.save(test_file)
subprocess.Popen([prog, test_file])